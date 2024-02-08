from django.contrib.auth.models import User
from django.contrib.auth import authenticate
from rest_framework import viewsets, permissions, status
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.authtoken.models import Token
from .models import Client, Courier, Order, Notification
from rest_framework import serializers
from .serializers import ClientSerializer, CourierSerializer, OrderSerializer, UserSerializer, NotificationSerializer
from .permissions import IsSuperuser, IsOwnerOrReadOnly, IsClientOrSuperuser, IsAdminUser
from .filters import OrderFilter
from rest_framework.decorators import action
from django.shortcuts import get_object_or_404
from django.db import transaction
from rest_framework.filters import OrderingFilter
from .utils import send_sms_via_smsoffice

from django.db.models.signals import post_save
from django.dispatch import receiver


from openpyxl import load_workbook
from io import BytesIO

class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAdminUser]

class ClientViewSet(viewsets.ModelViewSet):
    queryset = Client.objects.all()
    serializer_class = ClientSerializer
    permission_classes = [permissions.IsAdminUser]

class CourierViewSet(viewsets.ModelViewSet):
    queryset = Courier.objects.all()
    serializer_class = CourierSerializer
    permission_classes = [permissions.IsAdminUser]

class OrderViewSet(viewsets.ModelViewSet):
    queryset = Order.objects.all()
    serializer_class = OrderSerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_class = OrderFilter
    ordering_fields = '__all__'

    def get_queryset(self):
        user = self.request.user
        if user.is_superuser:
            return Order.objects.all()
        elif hasattr(user, 'client'):
            return Order.objects.filter(client=user.client)
        elif hasattr(user, 'courier'):
            return Order.objects.filter(courier=user.courier)
        else:
            return Order.objects.none()

    def get_permissions(self):
        if self.request.user.is_superuser:
            return [permissions.IsAuthenticated()]

        if self.action == 'destroy':
            return [IsSuperuser()]
        elif self.action in ['list', 'create', 'retrieve']:
            return [permissions.IsAuthenticated()]
        else:
            return [IsOwnerOrReadOnly()]

    def perform_create(self, serializer):
        if hasattr(self.request.user, 'client'):
            serializer.save(client=self.request.user.client)
        elif self.request.user.is_superuser:
            serializer.save()

    def perform_update(self, serializer):
        user = self.request.user
        order = serializer.instance

        if user.is_superuser:
            pass
        elif hasattr(user, 'client'):
            restricted_fields = ['item_price', 'courier_fee', 'sum', 'status', 'courier', 'client', 'status_approved', 'staged_status',]
            for field in restricted_fields:
                if field in serializer.validated_data:
                    serializer.validated_data.pop(field, None)
        elif hasattr(user, 'courier'):
            allowed_fields = ['comment', 'staged_status']
            for field in list(serializer.validated_data):
                if field not in allowed_fields:
                    serializer.validated_data.pop(field, None)

        if 'staged_status' in serializer.validated_data and order.status_approved:
            raise serializers.ValidationError({"detail": "სტატუსი ვერ შეიცვლება მას შემდეგ, რაც დადასტურდა ადმინისტრატორის მიერ!"})


        super().perform_update(serializer)



    def destroy(self, request, *args, **kwargs):
        order = get_object_or_404(Order, pk=kwargs.get('pk'))
        order.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    
    @action(detail=True, methods=['post'], permission_classes=[IsSuperuser])
    def approve_status(self, request, pk=None):
        order = self.get_object()
        order.status = order.staged_status
        order.status_approved = True
        order.save()
        return Response({"detail": "შეკვეთის სტატუსი დადასტურდა!"})

    @action(detail=True, methods=['post'], permission_classes=[IsSuperuser])
    def disapprove_status(self, request, pk=None):
        order = self.get_object()
        order.staged_status = 'DF'
        order.status_approved = False
        order.save()
        return Response({"detail": "შეკვეთის სტატუსი არ დადასტურდა!"})

    @action(methods=['delete'], detail=False, url_path='delete-batch', permission_classes=[IsSuperuser])
    def delete_batch(self, request, *args, **kwargs):

        if not request.user.is_superuser:
            return Response({'detail': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)

        order_ids = request.data.get('ids', [])
        if not order_ids:
            return Response({'detail': 'No order IDs provided'}, status=status.HTTP_400_BAD_REQUEST)

        Order.objects.filter(id__in=order_ids).delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    
    @action(detail=False, methods=['post'], permission_classes=[IsSuperuser])
    def set_order(self, request):
        order_data = request.data.get('order')

        if not order_data:
            return Response({'detail': 'No order data provided'}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            for index, order_id in enumerate(order_data):
                Order.objects.filter(id=order_id).update(order_position=index)

        return Response({'detail': 'შეკვეთა წარმატებით განახლდა!'}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['post'], permission_classes=[IsSuperuser])
    def send_bulk_sms(self, request):
        order_ids = request.data.get('order_ids')

        if not order_ids:
            return Response({"detail": "საჭიროა შეკვეთების კოდები!"}, status=status.HTTP_400_BAD_REQUEST)

        orders = Order.objects.filter(id__in=order_ids).values_list('phone_number', flat=True)
        destinations = ','.join(orders)
        
        response = send_sms_via_smsoffice(destinations)
        
        if response and response.get("Success"):
            return Response({"detail": "SMS წარმატებით გაიგზავნა."}, status=status.HTTP_200_OK)
        else:
            return Response({"detail": "შეცდომა SMS გაგზავნისას", "response": response}, status=status.HTTP_400_BAD_REQUEST)


    @action(detail=False, methods=['post'], permission_classes=[IsSuperuser])
    def assign_courier(self, request, *args, **kwargs):
        courier_id = request.data.get('courier_id')
        order_ids = request.data.get('order_ids', [])

        if not Courier.objects.filter(id=courier_id).exists():
            return Response({'detail': 'კურიერი არ არსებობს.'}, status=status.HTTP_404_NOT_FOUND)
        
        orders = Order.objects.filter(id__in=order_ids)

        updated_orders_count = orders.update(courier_id=courier_id)

        return Response({
            'detail': f'შეკვეთები: {updated_orders_count} წარმატებით გადაეცა კურიერს',
            'assigned_orders_count': updated_orders_count
        }, status=status.HTTP_200_OK)

class LoginView(APIView):
    def post(self, request, *args, **kwargs):
        username = request.data.get("username")
        password = request.data.get("password")
        user = authenticate(username=username, password=password)
        if user:
            token, created = Token.objects.get_or_create(user=user)

            user_data = {
                'id': user.id,
                'username': user.username,
                'is_staff': user.is_staff,
                'user_type': 'admin' if user.is_superuser else 'client' if hasattr(user, 'client') else 'courier' if hasattr(user, 'courier') else 'unknown'
            }

            if user.is_superuser:
                user_data['profile'] = {
                    'name': "Admin User",
                    'email': user.email
                }
            elif hasattr(user, 'client'):
                client = user.client
                user_data['profile'] = {
                    'id': client.id,
                    'name': client.name,
                    'representative_full_name': client.representative_full_name,
                    'email': client.email,
                    'phone_number': client.phone_number,
                    'addresses': client.addresses
                }
            elif hasattr(user, 'courier'):
                courier = user.courier
                user_data['profile'] = {
                    'id': courier.id,
                    'name': courier.name,
                    'email': user.email,
                    'phone_number': courier.phone_number
                }

            return Response({"token": token.key, "user_data": user_data}, status=status.HTTP_200_OK)
        
        return Response({"error": "არასწორი მონაცემები!"}, status=status.HTTP_400_BAD_REQUEST)

class NotificationViewSet(viewsets.ModelViewSet):
    queryset = Notification.objects.all()
    serializer_class = NotificationSerializer
    permission_classes=[IsSuperuser]

    def get_queryset(self):
        return Notification.objects.filter(admin_user=self.request.user).order_by('-created_at')

    def update(self, request, *args, **kwargs):
        notification = self.get_object()
        notification.is_read = True
        notification.save()
        return Response({'status': 'შეტყობინებები მონიშნულია როგორც წაკითხული'}, status=status.HTTP_200_OK)


class ExcelUploadView(APIView):
    permission_classes = [IsClientOrSuperuser, permissions.IsAuthenticated]

    def post(self, request, *args, **kwargs):
        excel_file = request.FILES.get('file')

        if not excel_file.name.endswith('.xlsx'):
            return Response({"error": "ფაილი არ არის ექსელის ფორმატში!"}, status=status.HTTP_400_BAD_REQUEST)

        workbook = load_workbook(filename=BytesIO(excel_file.read()))
        sheet = workbook.active
        with transaction.atomic():
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if request.user.is_superuser:
                    client = None
                else:
                    client = request.user.client

                order_data = {
                    'city': row[0],
                    'addressee_full_name': row[1],
                    'phone_number': row[2],
                    'address': row[3],
                    'comment': row[4],
                    'item_price': row[5],
                    'courier_fee': row[6],
                    'sum': row[7],
                    'client': client.pk if client else None,
                }

                serializer = OrderSerializer(data=order_data)
                if serializer.is_valid():
                    serializer.save()
                else:
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        return Response({"success": "ფაილი წარმატებით აიტვირთა!"}, status=status.HTTP_201_CREATED)